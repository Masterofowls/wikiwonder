"""Multi-format file import parsers."""
from __future__ import annotations

import csv
import io
import json
from pathlib import Path

from apps.imports.services import plain_text_to_markdown


def parse_uploaded_file(uploaded_file) -> dict:
    name = uploaded_file.name
    ext = Path(name).suffix.lower()
    raw_bytes = uploaded_file.read()

    if ext in {".md", ".markdown"}:
        text = raw_bytes.decode("utf-8", errors="replace")
        return {"title": Path(name).stem.replace("-", " ").title(), "markdown": text, "format": "md"}

    if ext == ".txt":
        text = raw_bytes.decode("utf-8", errors="replace")
        return {"title": Path(name).stem.replace("-", " ").title(), "markdown": plain_text_to_markdown(text), "format": "txt"}

    if ext == ".json":
        data = json.loads(raw_bytes.decode("utf-8"))
        if isinstance(data, dict) and "content" in data:
            return {
                "title": data.get("title", Path(name).stem),
                "markdown": data.get("content", ""),
                "format": "json",
            }
        return {"title": Path(name).stem, "markdown": f"```json\n{json.dumps(data, indent=2)}\n```", "format": "json"}

    if ext == ".csv":
        text = raw_bytes.decode("utf-8", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        rows = list(reader)
        md_lines = ["| " + " | ".join(reader.fieldnames or []) + " |"]
        if reader.fieldnames:
            md_lines.append("| " + " | ".join(["---"] * len(reader.fieldnames)) + " |")
        for row in rows:
            md_lines.append("| " + " | ".join(str(row.get(h, "")) for h in (reader.fieldnames or [])) + " |")
        return {"title": Path(name).stem, "markdown": "\n".join(md_lines), "format": "csv", "rows": len(rows)}

    if ext in {".xlsx", ".xls"}:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return {"title": Path(name).stem, "markdown": "", "format": "xlsx"}
        headers = [str(c or "") for c in rows[0]]
        md_lines = ["| " + " | ".join(headers) + " |", "| " + " | ".join(["---"] * len(headers)) + " |"]
        for row in rows[1:]:
            md_lines.append("| " + " | ".join(str(c or "") for c in row) + " |")
        return {"title": Path(name).stem, "markdown": "\n".join(md_lines), "format": "xlsx"}

    if ext == ".pdf":
        from pypdf import PdfReader

        reader = PdfReader(io.BytesIO(raw_bytes))
        parts = []
        for page in reader.pages[:10]:
            parts.append(page.extract_text() or "")
        text = "\n\n".join(parts)
        return {
            "title": Path(name).stem,
            "markdown": plain_text_to_markdown(text),
            "format": "pdf",
            "pages": len(reader.pages),
        }

    if ext in {".doc", ".docx"}:
        from docx import Document

        doc = Document(io.BytesIO(raw_bytes))
        text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
        return {"title": Path(name).stem, "markdown": plain_text_to_markdown(text), "format": "docx"}

    if ext == ".fb2":
        text = raw_bytes.decode("utf-8", errors="replace")
        import re

        title_match = re.search(r"<book-title>([^<]+)</book-title>", text)
        title = title_match.group(1) if title_match else Path(name).stem
        body = re.sub(r"<[^>]+>", " ", text)
        body = " ".join(body.split())[:8000]
        return {"title": title, "markdown": plain_text_to_markdown(body), "format": "fb2"}

    if ext in {".html", ".htm"}:
        html_text = raw_bytes.decode("utf-8", errors="replace")
        return {
            "title": Path(name).stem,
            "markdown": f'```html\n{html_text[:12000]}\n```',
            "format": "html",
        }

    text = raw_bytes.decode("utf-8", errors="replace")
    return {"title": Path(name).stem, "markdown": plain_text_to_markdown(text), "format": ext.lstrip(".") or "bin"}
